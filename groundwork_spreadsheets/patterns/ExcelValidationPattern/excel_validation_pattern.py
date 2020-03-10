"""
Groundwork Excel read/write routines using openpyxl
"""
import datetime
import json
import os
import re

import openpyxl
from groundwork.patterns.gw_base_pattern import GwBasePattern
from jsonschema import validate, ValidationError, SchemaError
from openpyxl.utils import get_column_letter

JSON_SCHEMA_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'assets', 'excel_config_schema.json')


class ExcelValidationPattern(GwBasePattern):
    """
    Groundwork Excel read/write routines using openpyxl
    """

    def __init__(self, *args, **kwargs):
        super(ExcelValidationPattern, self).__init__(*args, **kwargs)
        self.excel_validation = ExcelValidationPlugin(plugin=self)


class ExcelValidationPlugin:
    """
    Plugin level class of the Excel validation pattern.
    Note: There are no instances of the ExcelValidationPattern on the app.
    """

    def __init__(self, plugin):
        """
        :param plugin: The plugin, which wants to use documents
        :type plugin: GwBasePattern
        """
        self._plugin = plugin
        self._app = plugin.app
        self.excel_config = None

    def read_excel(self, excel_config_json_path, excel_workbook_path):
        """
        Main routine to read an Excel sheet.

        :param excel_config_json_path: The configuration json file
        :param excel_workbook_path: Relative or absolute path to an Excel workbook
        :return: Data dictionary with rows/colums as keys and a dictionary of "header": value as items
        """

        # The exceptions raised in this method shall be raised to the plugin level
        self.excel_config = self._validate_json(excel_config_json_path)

        #########################
        # Get index configuration
        #########################
        orientation = self.excel_config['orientation']
        header_idx_cfg_row_first = self.excel_config['headers_index_config']['row_index']['first']
        header_idx_cfg_row_last = self.excel_config['headers_index_config']['row_index']['last']
        header_idx_cfg_col_first = self.excel_config['headers_index_config']['column_index']['first']
        header_idx_cfg_col_last = self.excel_config['headers_index_config']['column_index']['last']
        data_idx_cfg_row_first = self.excel_config['data_index_config']['row_index']['first']
        data_idx_cfg_row_last = self.excel_config['data_index_config']['row_index']['last']
        data_idx_cfg_col_first = self.excel_config['data_index_config']['column_index']['first']
        data_idx_cfg_col_last = self.excel_config['data_index_config']['column_index']['last']

        ############################################################################
        # Rotate coordinates so we can work with virtually column based all the time
        # Just before addressing the cells a coordinate transformation is done again
        ############################################################################
        if orientation == 'column_based':
            # Correct the header and data matrix according to the orientation
            # We rotate the row_based layout so we can work with column_based in mind all the time
            # Name schema: corrected headers/data index config column/row first/last
            corr_header_idx_cfg_row_first = header_idx_cfg_row_first
            corr_header_idx_cfg_row_last = header_idx_cfg_row_last
            corr_header_idx_cfg_col_first = header_idx_cfg_col_first
            corr_header_idx_cfg_col_last = header_idx_cfg_col_last
            corr_data_idx_cfg_row_first = data_idx_cfg_row_first
            corr_data_idx_cfg_row_last = data_idx_cfg_row_last
            corr_data_idx_cfg_col_first = data_idx_cfg_col_first
            corr_data_idx_cfg_col_last = data_idx_cfg_col_last
        else:
            # row_based layout
            corr_header_idx_cfg_row_first = header_idx_cfg_col_first
            corr_header_idx_cfg_row_last = header_idx_cfg_col_last
            corr_header_idx_cfg_col_first = header_idx_cfg_row_first
            corr_header_idx_cfg_col_last = header_idx_cfg_row_last
            corr_data_idx_cfg_row_first = data_idx_cfg_col_first
            corr_data_idx_cfg_row_last = data_idx_cfg_col_last
            corr_data_idx_cfg_col_first = data_idx_cfg_row_first
            corr_data_idx_cfg_col_last = data_idx_cfg_row_last

        # Used for log and exception messages
        oriented_row_text = "row" if orientation == 'column_based' else "column"
        oriented_column_text = "column" if orientation == 'column_based' else "row"

        ######################################
        # Set defaults of headers_index_config
        ######################################
        if type(corr_header_idx_cfg_row_first) is not int:
            # Assume the user wants to use the first row as header row
            corr_header_idx_cfg_row_first = 1
            self._plugin.log.debug("Config update: Setting headers_index_config -> {0}_index -> first to 1".format(
                oriented_row_text))

        if type(corr_header_idx_cfg_row_last) is not int:
            # Only 1 header row is supported currently, set last header row to first header row
            corr_header_idx_cfg_row_last = corr_header_idx_cfg_row_first
            self._plugin.log.debug("Config update: Setting headers_index_config -> {0}_index -> last to {1}".format(
                oriented_row_text, corr_header_idx_cfg_row_first))

        if type(corr_header_idx_cfg_col_first) is not int:
            # Assume the user wants to start at the first column
            corr_header_idx_cfg_col_first = 1
            self._plugin.log.debug("Config update: Setting headers_index_config -> {0}_index -> first to 1".format(
                oriented_column_text))

        if type(corr_header_idx_cfg_col_last) is not int:
            if type(corr_data_idx_cfg_col_last) is int:
                # We don't have a last column in the header config,
                # so we use what we have in the data config
                corr_header_idx_cfg_col_last = corr_data_idx_cfg_col_last
                self._plugin.log.debug("Config update: Setting headers_index_config -> {0}_index -> first to "
                                       "{1}".format(oriented_column_text, corr_header_idx_cfg_col_last))

        ###################################
        # Set defaults of data_index_config
        ###################################
        if type(corr_data_idx_cfg_row_first) is not int:
            # Assume the first data row is the next after corr_header_idx_cfg_row_last
            corr_data_idx_cfg_row_first = corr_header_idx_cfg_row_last + 1
            self._plugin.log.debug("Config update: Setting data_index_config -> {0}_index -> first to {1}".format(
                oriented_row_text, corr_data_idx_cfg_row_first))

        # corr_data_idx_cfg_row_last has no defaults - the user input is master

        if type(corr_data_idx_cfg_col_first) is not int:
            # Assume the first data column is equal to the first header column
            corr_data_idx_cfg_col_first = corr_header_idx_cfg_col_first
            self._plugin.log.debug("Config update: Setting data_index_config -> {0}_index -> first to {1}".format(
                oriented_column_text, corr_data_idx_cfg_col_first))

        if type(corr_data_idx_cfg_col_last) is not int:
            if type(corr_header_idx_cfg_col_last) is int:
                # We don't have a last column in the data config,
                # so we use what we have in the header config
                corr_data_idx_cfg_col_last = corr_header_idx_cfg_col_last
                self._plugin.log.debug("Config update: Setting data_index_config -> {0}_index -> last to "
                                       "{1}".format(oriented_column_text, corr_data_idx_cfg_col_last))

        ################################
        # Some more logic checks on rows
        # (matrix size and row order)
        ################################
        if corr_header_idx_cfg_row_last != corr_header_idx_cfg_row_first:
            # We can compare both because at this point they have to be integer
            # Multi line headers given
            self._raise_value_error("Config error: Multi line (grouped) headers are not yet supported. "
                                    "First and last header {0} must be equal.".format(oriented_row_text))

        if corr_data_idx_cfg_row_first <= corr_header_idx_cfg_row_last:
            # We can compare both because at this point they have to be integer
            # The data section is above the header section
            self._raise_value_error("Config error: headers_index_config -> {0}_index -> last is greater than "
                                    "data_index_config -> {0}_index -> first.".format(oriented_row_text))

        if type(corr_data_idx_cfg_row_last) is int:
            if corr_data_idx_cfg_row_last < corr_data_idx_cfg_row_first:
                # The last data row is smaller than the first
                self._raise_value_error("Config error: data_index_config -> {0}_index -> first is greater than "
                                        "data_index_config -> {0}_index -> last.".format(oriented_row_text))

        ###################################
        # Some more logic checks on columns
        # (mismatches)
        ###################################
        if corr_header_idx_cfg_col_first != corr_data_idx_cfg_col_first:
            # We can compare both because at this point they have to be integer
            # First column mismatch
            self._raise_value_error("Config error: header_index_config -> {0}_index -> first is not equal to "
                                    "data_index_config -> {0}_index -> first.".format(oriented_column_text))

        if type(corr_header_idx_cfg_col_last) is int:
            if type(corr_data_idx_cfg_col_last) is int:
                if corr_header_idx_cfg_col_last != corr_data_idx_cfg_col_last:
                    # Last columns are given but do not match
                    self._raise_value_error(
                        "Config error: header_index_config -> {0}_index -> last ({1}) is not equal to "
                        "data_index_config -> {0}_index -> last ({2}).".format(
                            oriented_column_text,
                            corr_header_idx_cfg_col_last,
                            corr_data_idx_cfg_col_last
                        ))

        if type(corr_header_idx_cfg_col_last) is not int:
            if type(corr_data_idx_cfg_col_last) is not int:
                # The column count search is done on the header row, not on data
                if corr_data_idx_cfg_col_last != 'automatic':
                    self._raise_value_error(
                        "Config error: data_index_config -> {0}_index -> last ({1}) may only be an integer or "
                        "contain the value 'automatic'.".format(oriented_column_text,
                                                                corr_data_idx_cfg_col_last))

        #########################################
        # Set defaults for optional config values
        #########################################
        if 'sheet_config' not in self.excel_config:
            self.excel_config['sheet_config'] = 'active'

        for data_type_config in self.excel_config['data_type_config']:
            # default for possible problems should be strict if user tells nothing
            if 'fail_on_type_error' not in data_type_config:
                data_type_config['fail_on_type_error'] = True
            if 'fail_on_empty_cell' not in data_type_config:
                data_type_config['fail_on_empty_cell'] = True
            if 'fail_on_header_not_found' not in data_type_config:
                data_type_config['fail_on_header_not_found'] = True

            # default type is automatic
            if 'type' not in data_type_config:
                data_type_config['type'] = {'base': 'automatic'}

        # Set global filter properties
        # Defensive variant is True for all options
        if 'filter_properties' not in self.excel_config:
            self.excel_config['filter_properties'] = {}
        if 'excluded_fail_on_type_error' not in self.excel_config['filter_properties']:
            self.excel_config['filter_properties']['excluded_fail_on_type_error'] = True
        if 'excluded_fail_on_empty_cell' not in self.excel_config['filter_properties']:
            self.excel_config['filter_properties']['excluded_fail_on_empty_cell'] = True
        if 'excluded_enable_logging' not in self.excel_config['filter_properties']:
            self.excel_config['filter_properties']['excluded_enable_logging'] = True

        ############################
        # Get the workbook and sheet
        ############################
        wb = openpyxl.load_workbook(excel_workbook_path, data_only=True)
        ws = self._get_sheet(wb)

        #############################
        # Determine header row length
        #############################
        if type(corr_header_idx_cfg_col_last) == int:
            # corr_header_idx_cfg_col_last already has the final value
            pass
        elif corr_header_idx_cfg_col_last == 'automatic':
            # automatic: use the length of the header row
            corr_header_idx_cfg_col_last = len(
                ws[self._transform_coordinates(row=corr_header_idx_cfg_row_first)])
            self._plugin.log.debug("Config update: Last header {0} was set to {1} using the 'automatic' "
                                   "mechanism.".format(oriented_column_text, corr_header_idx_cfg_col_last))
        else:
            # severalEmptyCells chosen
            target_empty_cell_count = int(corr_header_idx_cfg_col_last.split(':')[1])
            empty_cell_count = 0
            curr_column = corr_header_idx_cfg_col_first
            while empty_cell_count < target_empty_cell_count:
                value = ws[self._transform_coordinates(corr_header_idx_cfg_row_first,
                                                       curr_column)].value
                if value is None:
                    empty_cell_count += 1
                curr_column += 1
            corr_header_idx_cfg_col_last = curr_column - target_empty_cell_count - 1
            self._plugin.log.debug("Config update: Last header {0} was set to {1} using the 'automatic' "
                                   "mechanism.".format(oriented_column_text, corr_header_idx_cfg_col_last))

        ###################################
        # Determine header column locations
        ###################################
        spreadsheet_headers2columns = {}
        for column in range(corr_header_idx_cfg_col_first, corr_header_idx_cfg_col_last + 1):
            value = ws[self._transform_coordinates(corr_header_idx_cfg_row_first, column)].value
            if value is not None:
                spreadsheet_headers2columns[value] = column
            else:
                # if the value is None we have either
                # - one or more empty header cell in between 2 filled header cells.
                # - some empty header cells at the end of the row.
                #   That might happen because when choosing 'automatic' header row detection, openpyxl functionality
                #   is used. It always returns the length of the longest row in the whole sheet.
                #   If that is not the header row, we have empty cells at the end of the header row.
                #   However, that is not a problem as header values 'None' are not added.
                pass
        spreadsheet_headers = spreadsheet_headers2columns.keys()

        ###############################################################
        # Check for not existing headers on spreadsheet and config side
        ###############################################################
        # Build a data_type_config dictionary with header as key
        config_header_dict = {x['header']: x for x in self.excel_config['data_type_config']}

        # Check: Are config data_type_config headers unique?
        if len(config_header_dict.keys()) != len(set(config_header_dict.keys())):
            self._raise_value_error("Config error: data_type_config -> header duplicate entries found.")

        # Check for configured headers not found in spreadsheet
        missing_headers_in_spreadsheet = list(set(config_header_dict.keys()) - set(spreadsheet_headers))
        for header in missing_headers_in_spreadsheet:
            # Check if the fail_on_header_not_found is true
            data_type = [x for x in self.excel_config['data_type_config'] if x['header'] == header][0]
            msg = u"Config error: The header '{0}' could not be found in the spreadsheet.".format(header)
            if data_type['fail_on_header_not_found']:
                self._raise_value_error(msg)
            else:
                self._plugin.log.error(msg)

        # Check for spreadsheet headers not found in config
        missing_headers_in_config = list(set(spreadsheet_headers) - set(config_header_dict.keys()))
        if missing_headers_in_config:
            self._plugin.log.debug(u"The following spreadsheet headers are not configured for reading: {0}".format(
                ', '.join([header.replace('\n', ' ') for header in missing_headers_in_config])))
            for header in missing_headers_in_config:
                del spreadsheet_headers2columns[header]

        #########################
        # Determine last data row
        #########################
        if type(corr_data_idx_cfg_row_last) == int:
            # do nothing, the value is already set
            pass
        elif corr_data_idx_cfg_row_last == 'automatic':
            # Take data from library
            corr_data_idx_cfg_row_last = corr_data_idx_cfg_row_first
            for curr_column in range(corr_header_idx_cfg_col_first,
                                     corr_header_idx_cfg_col_last):
                len_curr_column = len(ws[self._transform_coordinates(column=curr_column)])
                if len_curr_column > corr_data_idx_cfg_row_last:
                    corr_data_idx_cfg_row_last = len_curr_column
            self._plugin.log.debug("Config update: Last data {0} was set to {1} using the 'automatic' "
                                   "mechanism.".format(oriented_row_text, corr_data_idx_cfg_row_last))
        else:
            # severalEmptyCells is chosen
            target_empty_rows_count = int(corr_data_idx_cfg_row_last.split(':')[1])
            last_row_detected = False
            curr_row = corr_data_idx_cfg_row_first
            empty_rows_count = 0
            while not last_row_detected:
                # go through rows
                all_columns_empty = True
                for header in spreadsheet_headers2columns:
                    curr_column = spreadsheet_headers2columns[header]
                    value = ws[self._transform_coordinates(curr_row, curr_column)].value
                    if value is not None:
                        all_columns_empty = False
                        break
                if all_columns_empty:
                    empty_rows_count += 1
                if empty_rows_count >= target_empty_rows_count:
                    last_row_detected = True
                else:
                    curr_row += 1
            corr_data_idx_cfg_row_last = curr_row - target_empty_rows_count
            self._plugin.log.debug("Config update: Last data {0} was set to {1} using the 'severalEmptyCells' "
                                   "mechanism.".format(oriented_row_text, corr_data_idx_cfg_row_last))

        #################################################
        # Go through the rows, read and validate the data
        #################################################
        final_dict = {}
        for curr_row in range(corr_data_idx_cfg_row_first, corr_data_idx_cfg_row_last + 1):
            # Go through rows
            final_dict[curr_row] = {}

            msg_queue = {
                'fail_on_empty_cell': {
                    'exceptions': [],
                    'logs': []
                },
                'fail_on_type_error': {
                    'exceptions': [],
                    'logs': []
                }
            }
            is_row_excluded = False

            for header in spreadsheet_headers2columns:
                # Go through columns
                curr_column = spreadsheet_headers2columns[header]
                cell_index_str = self._transform_coordinates(curr_row, curr_column)
                value = ws[cell_index_str].value

                # Start the validation
                config_header = config_header_dict[header]

                if value is None:
                    msg = u"The '{0}' in cell {1} is empty".format(header, cell_index_str)
                    if config_header['fail_on_empty_cell']:
                        msg_queue['fail_on_empty_cell']['exceptions'].append(msg)
                    else:
                        msg_queue['fail_on_empty_cell']['logs'].append(msg)
                else:
                    if config_header['type']['base'] == 'automatic':
                        pass
                    elif config_header['type']['base'] == 'date':
                        if not isinstance(value, datetime.datetime):
                            msg = u'The value {0} in cell {1} is of type {2}; required by ' \
                                  u'specification is datetime'.format(value, cell_index_str, type(value))
                            if config_header['fail_on_type_error']:
                                msg_queue['fail_on_type_error']['exceptions'].append(msg)
                            else:
                                msg_queue['fail_on_type_error']['logs'].append(msg)
                    elif config_header['type']['base'] == 'enum':
                        filtered_enum_values = []
                        if 'filter' in config_header['type']:
                            filtered_enum_values = config_header['type']['filter']['whitelist_values']
                        if not self._is_string(value):
                            msg = u'The value {0} in cell {1} is of type {2}; required by ' \
                                  u'specification is a string type (enum)'.format(value, cell_index_str, type(value))
                            if config_header['fail_on_type_error']:
                                msg_queue['fail_on_type_error']['exceptions'].append(msg)
                            else:
                                msg_queue['fail_on_type_error']['logs'].append(msg)
                            if filtered_enum_values:
                                msg = 'Cannot apply enum filter to cell {1} because the type check failed.'
                                self._plugin.log.error(msg)
                        else:
                            valid_values = config_header['type']['enum_values']
                            if value not in valid_values:
                                msg = u'The value {0} in cell {1} is not contained in the given enum ' \
                                      u'[{2}]'.format(value, cell_index_str, ', '.join(valid_values))
                                if config_header['fail_on_type_error']:
                                    msg_queue['fail_on_type_error']['exceptions'].append(msg)
                                else:
                                    msg_queue['fail_on_type_error']['logs'].append(msg)
                                if filtered_enum_values:
                                    msg = 'Cannot apply enum filter to cell {1} because the enum values check failed.'
                                    self._plugin.log.error(msg)
                            else:
                                if filtered_enum_values and value not in filtered_enum_values:
                                    is_row_excluded = True
                                    self._plugin.log.debug(u"The {0} {1} was excluded due to an exclude filter on "
                                                           u"cell {2} ({3} not in [{4}]).".format(
                                                               oriented_row_text,
                                                               curr_row,
                                                               cell_index_str,
                                                               value,
                                                               ', '.join(filtered_enum_values)))

                    elif config_header['type']['base'] == 'float':
                        # TODO Allow int, too
                        if not isinstance(value, float):
                            msg = u'The value {0} in cell {1} is of type {2}; required by ' \
                                  u'specification is float'.format(value, cell_index_str, type(value))
                            if config_header['fail_on_type_error']:
                                msg_queue['fail_on_type_error']['exceptions'].append(msg)
                            else:
                                msg_queue['fail_on_type_error']['logs'].append(msg)
                        else:
                            if 'minimum' in config_header['type']:
                                if value < config_header['type']['minimum']:
                                    msg = u'The value {0} in cell {1} is smaller than the given minimum ' \
                                          u'of {2}'.format(value, cell_index_str, config_header['type']['minimum'])
                                    if config_header['fail_on_type_error']:
                                        msg_queue['fail_on_type_error']['exceptions'].append(msg)
                                    else:
                                        msg_queue['fail_on_type_error']['logs'].append(msg)
                            if 'maximum' in config_header['type']:
                                if value > config_header['type']['maximum']:
                                    msg = u'The value {0} in cell {1} is greater than the given maximum ' \
                                          u'of {2}'.format(value, cell_index_str, config_header['type']['maximum'])
                                    if config_header['fail_on_type_error']:
                                        msg_queue['fail_on_type_error']['exceptions'].append(msg)
                                    else:
                                        msg_queue['fail_on_type_error']['logs'].append(msg)
                    elif config_header['type']['base'] == 'integer':
                        # Integer values stored by Excel are returned as float (e.g. 3465.0)
                        # So we have to check if the float can be converted to int without precision loss
                        if isinstance(value, float):
                            if value.is_integer():
                                # the typecast to int may return int or long, depending on the size of value
                                value = int(value)
                        if self._is_type_int_long(value):
                            if 'minimum' in config_header['type']:
                                if value < config_header['type']['minimum']:
                                    msg = u'The value {0} in cell {1} is smaller than the given minimum ' \
                                          u'of {2}'.format(value, cell_index_str, config_header['type']['minimum'])
                                    if config_header['fail_on_type_error']:
                                        msg_queue['fail_on_type_error']['exceptions'].append(msg)
                                    else:
                                        msg_queue['fail_on_type_error']['logs'].append(msg)
                            if 'maximum' in config_header['type']:
                                if value > config_header['type']['maximum']:
                                    msg = u'The value {0} in cell {1} is greater than the given maximum ' \
                                          u'of {2}'.format(value, cell_index_str, config_header['type']['maximum'])
                                    if config_header['fail_on_type_error']:
                                        msg_queue['fail_on_type_error']['exceptions'].append(msg)
                                    else:
                                        msg_queue['fail_on_type_error']['logs'].append(msg)
                        else:
                            msg = u'The value {0} in cell {1} is of type {2}; required by ' \
                                  u'specification is int'.format(value, cell_index_str, type(value))
                            if config_header['fail_on_type_error']:
                                msg_queue['fail_on_type_error']['exceptions'].append(msg)
                            else:
                                msg_queue['fail_on_type_error']['logs'].append(msg)

                    elif config_header['type']['base'] == 'string':
                        if self._is_type_numeric(value):
                            convert_numbers = False
                            if 'convert_numbers' in config_header['type']:
                                convert_numbers = config_header['type']['convert_numbers']
                            if convert_numbers:
                                value = str(value)
                        if not self._is_string(value):
                            msg = u'The value {0} in cell {1} is of type {2}; required by ' \
                                  u'specification is string'.format(value, cell_index_str, type(value))
                            if config_header['fail_on_type_error']:
                                msg_queue['fail_on_type_error']['exceptions'].append(msg)
                            else:
                                msg_queue['fail_on_type_error']['logs'].append(msg)
                        else:
                            if 'pattern' in config_header['type']:
                                if re.search(config_header['type']['pattern'], value) is None:
                                    msg = u'The value {0} in cell {1} does not follow the ' \
                                          u'given pattern {2}'.format(value, cell_index_str,
                                                                      config_header['type']['pattern'])
                                    if config_header['fail_on_type_error']:
                                        msg_queue['fail_on_type_error']['exceptions'].append(msg)
                                    else:
                                        msg_queue['fail_on_type_error']['logs'].append(msg)

                final_dict[curr_row][header] = value

            if is_row_excluded:
                # All messages are either raised as exception or logged
                # If at all depends on the settings in self.excel_config['filter_properties']

                for msg_type, messages in msg_queue.items():
                    for msg in messages['exceptions']:
                        if self.excel_config['filter_properties']['excluded_' + msg_type]:
                            # This ends the program on the first exception message
                            self._raise_value_error(msg)
                        elif self.excel_config['filter_properties']['excluded_enable_logging']:
                            # In case we don't want to raise type errors as exception
                            # we log them in case the user configured so
                            self._plugin.log.warn(msg)
                    for msg in messages['logs']:
                        if self.excel_config['filter_properties']['excluded_enable_logging']:
                            self._plugin.log.warn(msg)
                del final_dict[curr_row]
            else:
                for msg_type, messages in msg_queue.items():
                    for msg in messages['exceptions']:
                        # This ends the program on the first exception message
                        self._raise_value_error(msg)
                    for msg in messages['logs']:
                        self._plugin.log.warn(msg)

        return final_dict

    @staticmethod
    def _is_string(value):
        """
        Tests if a value is of a string type
        Python2: unicode, str
        Python3: str

        :param value: Variable of any type
        :return: True if value is of string type else false
        """
        try:
            str_type = basestring
        except NameError:
            str_type = str
        return isinstance(value, str_type)

    @staticmethod
    def _is_type_numeric(value):
        """
        Tests if a value is of an numeric type.
        Python2: int, long, float
        Python3: int, float

        :param value: Variable of any type
        :return: True if value is of integer type else false
        """
        numeric_types = (int, float)
        try:
            long
            numeric_types = (int, long, float)
        except NameError:
            pass

        return isinstance(value, numeric_types)

    @staticmethod
    def _is_type_int_long(value):
        """
        Tests if a value is of an integer type.
        Python2: int, long
        Python3: int

        :param value: Variable of any type
        :return: True if value is of integer type else false
        """
        numeric_types = (int)
        try:
            long
            numeric_types = (int, long)
        except NameError:
            pass

        return isinstance(value, numeric_types)

    def _validate_json(self, excel_config_json_path):

        try:
            with open(excel_config_json_path, encoding='utf-8') as file_pointer:
                json_obj = json.load(file_pointer)
        # the file is not deserializable as a json object
        except ValueError as exc:
            self._plugin.log.error('Malformed JSON file: {0} \n {1}'.format(excel_config_json_path, exc))
            raise exc
        # some os error occured (e.g file not found or malformed path string)
        # have to catch two exception classes: in py2 : IOError; py3: OSError
        except (IOError, OSError) as exc:
            self._plugin.log.error(exc)
            # raise only OSError to make error handling in caller easier
            raise OSError()

        # validate json object if schema file path is there; otherwise throw warning
        try:
            with open(JSON_SCHEMA_FILE_PATH) as file_pointer:
                schema_obj = json.load(file_pointer)
        # the file is not deserializable as a json object
        except ValueError as exc:
            self._plugin.log.error('Malformed JSON schema file: {0} \n {1}'.format(JSON_SCHEMA_FILE_PATH, exc))
            raise exc
        # some os error occured (e.g file not found or malformed path string)
        # have to catch two exception classes:  in py2 : IOError; py3: OSError
        except (IOError, OSError) as exc:
            self._plugin.log.error(exc)
            # raise only OSError to make error handling in caller easier
            raise OSError()

        # do the validation
        try:
            validate(json_obj, schema_obj)
        except ValidationError as error:
            self._plugin.log.error("Validation failed: {0}".format(error.message))
            raise
        except SchemaError:
            self._plugin.log.error("Invalid schema file: {0}".format(JSON_SCHEMA_FILE_PATH))
            raise

        return json_obj

    def _get_sheet(self, workbook):

        # get sheet
        worksheet = None
        if type(self.excel_config['sheet_config']) == int:
            worksheet = workbook.worksheets[self.excel_config['sheet_config'] - 1]
        elif self.excel_config['sheet_config'] == 'active':
            worksheet = workbook.active
        elif self.excel_config['sheet_config'].startswith('name'):
            worksheet = workbook[self.excel_config['sheet_config'].split(':')[1]]
        elif self.excel_config['sheet_config'] == 'first':
            worksheet = workbook.worksheets[0]
        elif self.excel_config['sheet_config'] == 'last':
            worksheet = workbook.worksheets[len(workbook.get_sheet_names()) - 1]
        else:
            # This cannot happen if json validation was ok
            pass
        return worksheet

    def _raise_value_error(self, msg):
        self._plugin.log.error(msg)
        raise ValueError(msg.encode('UTF-8', 'ignore'))

    def _transform_coordinates(self, row=None, column=None):
        if row is None and column is None:
            raise ValueError("_transform_coordinates: row and column cannot both be None.")
        target_str = ''
        if self.excel_config['orientation'] == 'column_based':
            if column is not None:
                target_str = get_column_letter(column)
            if row is not None:
                target_str += str(row)
        else:
            if row is not None:
                target_str = get_column_letter(row)
            if column is not None:
                target_str += str(column)
        return target_str
