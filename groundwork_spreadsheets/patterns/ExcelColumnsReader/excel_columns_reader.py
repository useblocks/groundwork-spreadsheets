"""
Groundwork Excel read/write routines using openpyxl
"""

import logging

from groundwork.patterns.gw_base_pattern import GwBasePattern

import json
from jsonschema import validate, ValidationError, SchemaError
import openpyxl

json_file_path = "excel_config.json"
json_schema_file_path = "excel_config_schema.json"

try:
    with open(json_file_path) as f:
        json_obj = json.load(f)
# the file is not deserializable as a json object
except ValueError as e:
    print('Malformed JSON file: {0} \n {1}'.format(json_file_path, e))
    raise e
# some os error occured (e.g file not found or malformed path string)
# have to catch two exception classes: in py2 : IOError; py3: OSError
except (IOError, OSError) as e:
    print(e)
    # raise only OSError to make error handling in caller easier
    raise OSError()

# validate json object if schema file path is there; otherwise throw warning
try:
    with open(json_schema_file_path) as f:
        schema_obj = json.load(f)
# the file is not deserializable as a json object
except ValueError as e:
    print('Malformed JSON schema file: {0} \n {1}'.format(json_schema_file_path, e))
    raise e
# some os error occured (e.g file not found or malformed path string)
# have to catch two exception classes:  in py2 : IOError; py3: OSError
except (IOError, OSError) as e:
    print(e)
    # raise only OSError to make error handling in caller easier
    raise OSError()

# do the validation
try:
    validate(json_obj, schema_obj)
except ValidationError as error:
    print("Validation failed: {0}".format(error.message))
    raise
except SchemaError:
    print("Invalid schema file: {0}".format(json_schema_file_path))
    raise

wb = openpyxl.load_workbook('test.xlsx', data_only=True)
ws = wb.active
print(ws.max_row)
print(ws.max_column)
print(ws.dimensions)

print(ws['A2'].value)
print(type(ws['A2'].value))
print(ws['B2'].value)
print(type(ws['B2'].value))
print(wb.get_sheet_names())
